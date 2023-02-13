import pandas as pd
import openpyxl
import numpy as np
import re

sheet_location = "D:\coding\goodstudy\학생 관리표.xlsx"
wb = openpyxl.load_workbook(sheet_location, data_only=True)
sheet_list = wb.sheetnames
w1 = wb[sheet_list[0]]

def extract_days(str):
    pattern = "([0-9]+) - ([0-9]+)day"

    result = re.search(pattern, str)
    first_day = int(result.group(1))
    last_day = int(result.group(2))

    for day in range(first_day, last_day+1):
        f = open("D:/coding/goodstudy/단어/하이퍼2000/정답지/{}day.txt".format(day), "r")
        print(f.readline())
        f.close()

print(extract_days("11 - 12day"))
def name_list():
    name_list = []
    

    for i in range(2,17):
        name_list.append(w1.cell(row=i, column=2).value)

    return name_list

def show_term(input_name):
    namelist = name_list()
    if(input_name in namelist):
        name_index = namelist.index(input_name) + 1

    book_name = w1.cell(row=name_index, column=6).value
    if book_name == None:
        print_sentense = "단어시험이 없습니다."
        return print_sentense
    else:
        test_day = w1.cell(row=name_index, column=7).value
        print_sentense = "단어책 : {0}\n시험범위 : {1}".format(book_name,test_day)
        return print_sentense
        

    