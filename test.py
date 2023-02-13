import main
import openpyxl

sheet_location = "D:\coding\goodstudy\학생 관리표.xlsx"
names_list = main.name_list()
wb = openpyxl.load_workbook(sheet_location, data_only=True)
sheet_list = wb.sheetnames
w1 = wb[sheet_list[0]]

input_name = input("이름을 입력하세요 : ")
print("-" * 20)
if(input_name in names_list):
    name_index = names_list.index(input_name) + 1

else:
    print("입력한 이름이 존재하지 않습니다.")
    quit()

book_name = w1.cell(row=name_index, column=6).value
if book_name == None:
    print("단어시험이 없습니다.")
    quit()
else:
    test_day = w1.cell(row=name_index, column=7).value
    print(input_name + "의 오늘 단어시험 범위는",book_name,test_day + "입니다.")

main.extract_days(test_day)